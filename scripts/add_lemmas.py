#!/usr/bin/env python3
import sys
import argparse
import openpyxl
import xml.etree.ElementTree as ET

def load_token_annotations(path, sheet_name):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise SystemExit(f"Sheet '{sheet_name}' not found in {path}")
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}
    header = [ (c or '').strip().lower() for c in rows[0] ]
    try:
        id_idx = header.index('id')
        lemma_idx = header.index('lemma')
        tag_idx = header.index('tag')
    except ValueError as e:
        raise SystemExit(f"Required columns 'id', 'lemma', and 'tag' not found in header: {e}")
    mapping = {}
    for row in rows[1:]:
        if not row:
            continue
        tokid = row[id_idx]
        lemma = row[lemma_idx]
        tag = row[tag_idx]
        if tokid:
            mapping[str(tokid)] = {'lemma': str(lemma) if lemma else None, 'tag': str(tag) if tag else None}
    return mapping

def enrich_xml_with_annotations(xml_in, xml_out, mapping):
    tree = ET.parse(xml_in)
    root = tree.getroot()
    updated = 0
    for tok in root.iter('tok'):
        tid = tok.get('id')
        if tid and tid in mapping:
            data = mapping[tid]
            if data.get('lemma'):
                tok.set('lemma', data['lemma'])
            if data.get('tag'):
                tok.set('mfs', data['tag'])
            updated += 1
    tree.write(xml_out, encoding='utf-8', xml_declaration=True)
    return updated

def main():
    parser = argparse.ArgumentParser(description='Add lemma and tag annotations from Excel to XML tokens')
    parser.add_argument('tagged_file', nargs='?', default='data/HdE_DCE.tagged.xlsx', help='Path to the tagged Excel file (.xlsx) (default: data/HdE_DCE.tagged.xlsx)')
    parser.add_argument('input_xml', nargs='?', default='data/HdE_DCE.xml', help='Path to the input XML file (default: HdE_DCE.xml)')
    parser.add_argument('output_xml', nargs='?', default='HdE_DCE.enriched.xml', help='Path to the output XML file (default: HdE_DCE.enriched.xml)')
    parser.add_argument('--sheet', default='Revised', help='Sheet name in Excel file (default: Revised)')
    args = parser.parse_args()
    
    mapping = load_token_annotations(args.tagged_file, args.sheet)
    print(f'Loaded {len(mapping)} token annotations from {args.tagged_file}#{args.sheet}')
    updated = enrich_xml_with_annotations(args.input_xml, args.output_xml, mapping)
    print(f'Updated {updated} tokens; wrote {args.output_xml}')

if __name__ == '__main__':
    main()